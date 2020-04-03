#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Copyright (c) 2020 Cisco and/or its affiliates.

This software is licensed to you under the terms of the Cisco Sample
Code License, Version 1.1 (the "License"). You may obtain a copy of the
License at

               https://developer.cisco.com/docs/licenses

All use of the material herein must be in accordance with the terms of
the License. All rights not expressly granted by the License are
reserved. Unless required by applicable law or agreed to separately in
writing, software distributed under the License is distributed on an "AS
IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express
or implied.

Author: Jean-Pierre Muzard <jmuzard@cisco.com>
Created: April 2, 2020

This script reads an excel file, name is: ravpn_users.xlsx
It includes user accounts to be added or removed in asa configuration

asa_ravpn_users.py -h
usage: asa_ravpn-users.py [-h] customer device

Configure Remote access VPN users on ASA

positional arguments:
  customer    customer name
  device      Device name

optional arguments:
  -h, --help  show this help message and exit

Example:

python3 asa_ravpn_users.py customer1 avav1

"""
import argparse
import openpyxl
from MSPO_utils import get_MSPO_env, netmiko_connect
# Open ravpn_users.xlsx excel file
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Configure Remote access VPN users on ASA"
    )
    parser.add_argument('customer', help="customer name")
    parser.add_argument('device', help="Device name")
    args = parser.parse_args()
    cfg = get_MSPO_env()
    ssh_conn = netmiko_connect(args.customer, args.device)
# process excel file
    EXCEL_FILE = cfg.get('customer_vars') + "/"+  args.customer + "/sites/" + args.device + "/ravpn_users.xlsx"
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb['RAVPNusers']
    max_row = sheet.max_row + 1
    for r in range(2, max_row):
        user = sheet.cell(row=r, column=1).value
        password = sheet.cell(row=r, column=2).value
        action = sheet.cell(row=r, column=3).value
        if action == "add":
            asa_cfg_cmd = "username " + user + " password " + password + "\n"
            ssh_conn.send_config_set([asa_cfg_cmd])
            print("user: ", user, "configured")
        else:
            asa_cfg_cmd = "no username " + user
            ssh_conn.send_config_set([asa_cfg_cmd])
            print("user: ", user, "removed")
    ssh_conn.disconnect()
